import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import calendar
import os
from typing import List, Dict

class ExcelGenerator:
    def __init__(self):
        self.data = []
        self.output_folder = None
        
    def generate_excel_files(self, data: List[Dict], output_folder: str, dieta_amount: float = 0.00):
        """Generate Excel files for each employee"""
        self.data = data
        self.output_folder = output_folder
        
        # Group data by employee
        employees = {}
        for row in data:
            name = row['nombre']
            if name not in employees:
                employees[name] = []
            employees[name].append(row)
        
        generated_files = []
        
        for employee_name, employee_data in employees.items():
            file_path = self.create_or_update_employee_excel(employee_name, employee_data, dieta_amount)
            if file_path:
                generated_files.append(file_path)
        
        return generated_files
    
    def create_or_update_employee_excel(self, employee_name: str, employee_data: List[Dict], dieta_amount: float):
        """Create or update Excel file for a specific employee"""
        try:
            # Get month and year from first record
            if employee_data:
                first_date = datetime.strptime(employee_data[0]['fecha'], '%d/%m/%Y')
                month_name = first_date.strftime('%B').upper()
                year = first_date.year
                month_num = first_date.month
            else:
                return None
            
            # Define file path
            employee_name = employee_name.upper()
            
            filename = f"{employee_name.replace(' ', '_')}_GASTOS_{year}.xlsx"
            file_path = os.path.join(self.output_folder, filename)
            
            # Check if file exists
            if os.path.exists(file_path):
                print(f"Archivo existente encontrado para {employee_name}, agregando nueva hoja...")
                wb = load_workbook(file_path)
                
                # Check if sheet with month name already exists
                sheet_name = f"{month_name}_{year}"
                if sheet_name in wb.sheetnames:
                    print(f"Hoja {sheet_name} ya existe, sobrescribiendo...")
                    wb.remove(wb[sheet_name])
                
                # Create new worksheet
                ws = wb.create_sheet(title=sheet_name)
            else:
                print(f"Creando nuevo archivo para {employee_name}...")
                wb = Workbook()
                ws = wb.active
                # Name the first sheet with month and year
                sheet_name = f"{month_name}_{year}"
                ws.title = sheet_name
            
            # Fill the worksheet with data
            self.fill_worksheet(ws, employee_name, employee_data, dieta_amount, month_name, year, month_num)
            
            # Apply formatting
            self.apply_formatting(ws)
            
            # Configure page setup for A4 format
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0  # 0 = automático
            
            # Configure page margins (1 cm = 0.393701 inches)
            ws.page_margins = PageMargins(
                left=0.393701,   # 1 cm
                right=0.393701,  # 1 cm
                top=0.393701,    # 1 cm
                bottom=0.393701, # 1 cm
                header=0,        # Sin encabezado
                footer=0         # Sin pie de página
            )
            
            # Save file
            wb.save(file_path)
            print(f"Archivo guardado: {file_path}")
            
            return file_path
            
        except Exception as e:
            print(f"Error creating/updating Excel for {employee_name}: {e}")
            return None
    
    def fill_worksheet(self, ws, employee_name: str, employee_data: List[Dict], dieta_amount: float, month_name: str, year: int, month_num: int):
        """Fill worksheet with employee data"""
        
        # Header section
        ws['A1'] = 'MES'
        ws['C1'] = month_name
        ws['E1'] = 'AÑO'
        ws['F1'] = year
        
        ws['A2'] = 'NOMBRE'
        ws['C2'] = employee_name
        ws['E2'] = 'VALOR'
        ws['F2'] = dieta_amount
        
        # Table headers (row 5)
        headers = ['Fecha', 'DIETAS', 'ALOJAMIENTO', 'GASOIL', 'KM', 'PEAJE', 'OBRA', 'FIRMA']
        for col, header in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=header)
        
        # Get days in month
        days_in_month = calendar.monthrange(year, month_num)[1]
        
        # Create work days set from employee data
        work_days = set()
        projects_by_day = {}
        
        for record in employee_data:
            day = datetime.strptime(record['fecha'], '%d/%m/%Y').day
            work_days.add(day)
            # Extraer solo el código del proyecto (antes del primer guión)
            project_full = record['proyecto']
            if ' - ' in project_full:
                project_code = project_full.split(' - ')[0].strip()
            else:
                project_code = project_full.strip()
            projects_by_day[day] = project_code
        
        # Fill days (rows 6 to 36, exactamente como solicitas)
        for day in range(1, 32):  # Días 1 a 31 (filas 6 a 36)
            row_num = day + 5  # día 1 = fila 6, día 31 = fila 36
            ws.cell(row=row_num, column=1, value=day)  # Date
            
            for day in range(1, 32):  # Días 1 a 31 (filas 6 a 36)
                row_num = day + 5  # día 1 = fila 6, día 31 = fila 36
                ws.cell(row=row_num, column=1, value=day)  # Date
                
                if day in work_days:
                    # Para días de trabajo: B=F2, C,D,E,F=0, G=proyecto
                    ws.cell(row=row_num, column=2, value='=$F$2')  # DIETAS = F2
                    ws.cell(row=row_num, column=3, value=0)        # ALOJAMIENTO = 0
                    ws.cell(row=row_num, column=4, value=0)        # GASOIL = 0
                    ws.cell(row=row_num, column=5, value=0)        # KM = 0
                    ws.cell(row=row_num, column=6, value=0)        # PEAJE = 0
                    ws.cell(row=row_num, column=7, value=projects_by_day.get(day, ''))  # OBRA
                else:
                    # Para días sin trabajo: todas las columnas con fórmulas pero valores 0
                    ws.cell(row=row_num, column=7, value='')
        # Total row (row 37)
        total_row = 37
        ws.cell(row=total_row, column=1, value='TOTAL')
        
        # Totales con fórmulas de suma de B6:B36, C6:C36, etc.
        ws.cell(row=total_row, column=2, value='=SUM(B6:B36)')  # DIETAS
        ws.cell(row=total_row, column=3, value='=SUM(C6:C36)')  # ALOJAMIENTO
        ws.cell(row=total_row, column=4, value='=SUM(D6:D36)')  # GASOIL
        ws.cell(row=total_row, column=5, value='=SUM(E6:E36)')  # KM
        ws.cell(row=total_row, column=6, value='=SUM(F6:F36)')  # PEAJE
        
        # Gran total en E39 (suma de B37:F37)
        grand_total_row = 39
        ws.cell(row=grand_total_row, column=4, value='Total:')
        ws.cell(row=grand_total_row, column=5, value=f'=SUM(B{total_row}:F{total_row})')
        
        # Footer sections
        received_row = grand_total_row + 2
        ws.cell(row=received_row, column=1, value='Recibí:')
        
        date_row = received_row + 6
        ws.cell(row=date_row, column=1, value='Fecha:')
        
        # Declaration text with wrap text
        declaration_row = date_row + 2
        declaration_text = ("El trabajador declara que los días señalados en cada una de las fechas consignadas en este documento "
                          "realizó todas y cada una de las rutas y desplazamientos, e incurrió en los gastos que se indican para "
                          "cada una de ellas, en ejercicio de su cargo y/o actividad laboral, validando con esta firma la totalidad "
                          "de las mismas. Y para que conste firma el presente documento en la fecha señalada.")
        
        # Configure the declaration cell with wrap text
        declaration_cell = ws.cell(row=declaration_row, column=1, value=declaration_text)
        declaration_cell.alignment = Alignment(
            horizontal='justify',
            vertical='top',
            wrap_text=True
        )
        
        # Merge cells for declaration text
        ws.merge_cells(f'A{declaration_row}:H{declaration_row + 3}')
        
        # Set row height for declaration to accommodate text
        for row_offset in range(4):  # 4 filas para el texto
            ws.row_dimensions[declaration_row + row_offset].height = 30
    
    def apply_formatting(self, ws):
        """Apply formatting to the worksheet"""
        # Bold headers
        bold_font = Font(bold=True)
        
        # Header formatting
        for cell in ['A1', 'C1', 'E1', 'F1', 'A2', 'C2', 'E2', 'F2']:
            ws[cell].font = bold_font
        
        # Table headers formatting
        for col in range(1, 9):
            cell = ws.cell(row=5, column=col)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center')
        
        # Border for table
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to data table (rows 5 to 37)
        for row in range(5, 38):  # Incluye hasta row 37
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border
        
        # Center alignment for date column
        for row in range(6, 37):  # B6 a B36
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        # Right alignment for monetary values
        for row in range(6, 38):  # Incluye hasta row 37
            for col in range(2, 7):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        
        # Bold formatting for total rows
        for col in range(1, 9):
            ws.cell(row=37, column=col).font = bold_font  # Total row
        
        ws.cell(row=39, column=4).font = bold_font  # "Total:" label
        ws.cell(row=39, column=5).font = bold_font  # Grand total value
        
        # Adjust column widths
        column_widths = [8, 12, 15, 12, 8, 12, 30, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width